import json
import sys
import tempfile
import unittest
import zipfile
from pathlib import Path

import cv2
import numpy as np
import openpyxl
import pandas as pd


REPO_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(REPO_ROOT / "src"))

from edlsrobot.datasets.export_lerobot_episode_to_excel import export_episode_to_excel


def _write_test_video(path: Path, frames: list[np.ndarray], fps: int = 10) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    height, width = frames[0].shape[:2]
    writer = cv2.VideoWriter(
        str(path),
        cv2.VideoWriter_fourcc(*"mp4v"),
        fps,
        (width, height),
    )
    if not writer.isOpened():
        raise RuntimeError(f"Failed to create test video at {path}")

    for frame in frames:
        writer.write(frame)
    writer.release()


class LerobotEpisodeToExcelTest(unittest.TestCase):
    def test_export_episode_to_excel_flattens_rows_and_embeds_all_camera_frames(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            root = Path(tmp_dir) / "demo_dataset"
            data_dir = root / "data" / "chunk-000"
            meta_dir = root / "meta"
            videos_dir = root / "videos" / "chunk-000"
            episode_path = data_dir / "episode_000000.parquet"
            output_path = root / "episode_000000.xlsx"

            data_dir.mkdir(parents=True, exist_ok=True)
            meta_dir.mkdir(parents=True, exist_ok=True)

            info = {
                "video_path": "videos/chunk-{episode_chunk:03d}/{video_key}/episode_{episode_index:06d}.mp4",
                "features": {
                    "observation.state": {"dtype": "float32", "shape": [2]},
                    "action": {"dtype": "float32", "shape": [2]},
                    "observation.images.head": {"dtype": "video", "shape": [16, 24, 3]},
                    "observation.images.left_wrist": {"dtype": "video", "shape": [16, 24, 3]},
                    "frame_index": {"dtype": "int64", "shape": [1]},
                    "episode_index": {"dtype": "int64", "shape": [1]},
                    "timestamp": {"dtype": "float32", "shape": [1]},
                },
            }
            (meta_dir / "info.json").write_text(json.dumps(info), encoding="utf-8")

            df = pd.DataFrame(
                {
                    "observation.state": [
                        np.array([1.0, 2.0], dtype=np.float32),
                        np.array([3.0, 4.0], dtype=np.float32),
                    ],
                    "action": [
                        np.array([0.1, 0.2], dtype=np.float32),
                        np.array([0.3, 0.4], dtype=np.float32),
                    ],
                    "frame_index": [0, 1],
                    "episode_index": [0, 0],
                    "timestamp": [0.0, 0.1],
                }
            )
            df.to_parquet(episode_path, index=False)

            blue = np.zeros((16, 24, 3), dtype=np.uint8)
            blue[..., 2] = 255
            green = np.zeros((16, 24, 3), dtype=np.uint8)
            green[..., 1] = 255
            red = np.zeros((16, 24, 3), dtype=np.uint8)
            red[..., 0] = 255
            white = np.full((16, 24, 3), 255, dtype=np.uint8)

            _write_test_video(
                videos_dir / "observation.images.head" / "episode_000000.mp4",
                [blue, green],
            )
            _write_test_video(
                videos_dir / "observation.images.left_wrist" / "episode_000000.mp4",
                [red, white],
            )

            export_episode_to_excel(episode_path, output_path)

            self.assertTrue(output_path.exists())

            workbook = openpyxl.load_workbook(output_path)
            sheet = workbook.active
            headers = [cell.value for cell in sheet[1]]

            self.assertIn("observation.state", headers)
            self.assertIn("action", headers)
            self.assertIn("observation.images.head", headers)
            self.assertIn("observation.images.left_wrist", headers)

            header_index = {name: idx + 1 for idx, name in enumerate(headers)}
            self.assertEqual(sheet.cell(row=2, column=header_index["observation.state"]).value, "[1.0, 2.0]")
            self.assertEqual(sheet.cell(row=3, column=header_index["action"]).value, "[0.30000001192092896, 0.4000000059604645]")

            self.assertIn("README", workbook.sheetnames)
            readme = workbook["README"]
            self.assertEqual(readme["A1"].value, "action dimension")
            self.assertEqual(readme["B2"].value, "left_waist")
            self.assertEqual(readme["A15"].value, "action[13]")

            with zipfile.ZipFile(output_path) as archive:
                media_files = [name for name in archive.namelist() if name.startswith("xl/media/")]
            self.assertEqual(len(media_files), 4)


if __name__ == "__main__":
    unittest.main()
